# validate src.strategy against instructor's HO reference outputs.
# runs the engine on HO-5minHLV.csv for each (L, S) in the xlsx grids and compares Profit / WorstDD / StDev / #trades for IS and OS windows.
#
# Usage:  .venv/bin/python scripts/validate_ho.py

import os
import sys
import time
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.strategy import (
    load_hlv_csv,
    index_at_or_after,
    index_at_or_before,
    run_strategy,
)

DATA = ROOT / "instruction_and_data" / "HO-5minHLV.csv"
OPT_XLSX = ROOT / "instruction_and_data" / "HO Optimization.xlsx"
TEST_XLSX = ROOT / "instruction_and_data" / "HO MatLab Test.xlsx"

# HO parameters. PV=42000 matches TF Data col H=420 *100
# Slpg: main.m hard-codes 47; TF Data col V says 70.2. We use 47 here because the xlsx reference files were clearly generated with 47 (switching to 70.2
# moves our PnL *further* from the reference, not closer -- see notes in README). For the final rolling-walk-forward production run, switch to 70.2 per the PDF.
PV = 42_000.0
SLPG = 47.0
BARS_BACK = 17001

E0 = 100000.0

IS_START = "1980-01-01"
IS_END = "2000-01-01"
OS_START = "2000-01-01"
OS_END = "2023-03-23"  # main.m outSample(2) = 03/23/2023


def load_reference_grid(xlsx_path):
    # parse {(L, S): {'is': (profit, worst_dd, stdev, trades), 'os': (...)}}
    # column layout (1-based):
    #   B: header / stoppct string
    #   E..H: IS Profit, IS WorstDD, IS StDev, IS #trades
    #   I..L: OS Profit, OS WorstDD, OS StDev, OS #trades
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]
    
    out = {}
    current_L = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row:
            continue
        if row[1] == "calculating" and row[3] == "Length":
            current_L = int(row[5])
            continue
        tag = row[1]
        if isinstance(tag, str) and tag.endswith(":") and current_L is not None:
            try:
                s = float(tag.rstrip(":"))
            except ValueError:
                continue
            key = (current_L, round(s, 3))
            if row[2] == "in/out:":
                
                pis, dis, sis, tis = row[3], row[4], row[5], row[6]
                pos, dos, sos, tos = row[8], row[9], row[10], row[11]
                
            elif row[2] == "in":
                pis, dis, sis, tis = row[4], row[5], row[6], row[7]
                pos, dos, sos, tos = row[8], row[9], row[10], row[11]
            else:
                continue
            if pis is None:
                continue
            out[key] = {
                "is": (float(pis), float(dis), float(sis), float(tis)),
                "os": (float(pos), float(dos), float(sos), float(tos)),
            }
    return out


def fmt(x):
    return f"{x:12.4f}"


def check(label, got, want, tol_abs=0.05, tol_rel=1e-4):
    gp, gd, gs, gt = got
    wp, wd, ws_, wt = want
    
    pairs = [("Profit", gp, wp), ("WorstDD", gd, wd), ("StDev", gs, ws_), ("Trades", gt, wt)]
    ok = True
    print(f'  {label}:')
    for name, g, w in pairs:
        diff = g - w
        rel = abs(diff) / max(abs(w), 1e-9)
        this_ok = abs(diff) <= tol_abs or rel <= tol_rel
        ok = ok and this_ok
        mark = '' if this_ok else 'MISMATCH???'
        print(f'{name} got={fmt(g)} want={fmt(w)} diff={fmt(diff)} (rel {rel:.2e}){mark}')
    return ok


def main():
    t0 = time.time()
    print(f"Loading {DATA.name} ...")
    dt, H, L, C = load_hlv_csv(str(DATA))
    print(f"  {len(H):,} bars  {dt[0]} -> {dt[-1]}   ({time.time()-t0:.1f}s)")

    # matlab indInSample1 = max(sum(numTime<IS_START)+1, barsBack), 1-based.
    # BARS_BACK (=17001) in matlab 1-based == BARS_BACK-1 in python 0-based.
    is_lo = max(index_at_or_after(dt, IS_START), BARS_BACK - 1)
    is_hi = max(index_at_or_before(dt, IS_END), BARS_BACK - 1)
    os_lo = max(index_at_or_after(dt, OS_START), BARS_BACK - 1)
    os_hi = max(index_at_or_before(dt, OS_END), BARS_BACK - 1)

    print(f"IS bars [{is_lo:>7}, {is_hi:>7}]  ({dt[is_lo]} -> {dt[is_hi]})")
    print(f"OS bars [{os_lo:>7}, {os_hi:>7}]  ({dt[os_lo]} -> {dt[os_hi]})")

    # Two reference tables:
    #   HO Optimization.xlsx -- full grid. IS 1980-2000, OS 2000-2023.
    #   HO MatLab Test.xlsx  -- small grid. IS window same; OS is a short window (~10 trades), so we only compare IS here.
    # At common (L, S) points the two tables disagree on IS profit by ~10% while IS trade counts differ by 1-4 -- i.e. same decisions, different fill/slpg rules. Running both tells us which version our engine matches.
    ref_opt = load_reference_grid(OPT_XLSX)
    ref_test = load_reference_grid(TEST_XLSX)
    print(f"  HO Optimization: {len(ref_opt)} (L,S) pairs")
    print(f"  HO MatLab Test:  {len(ref_test)} (L,S) pairs")

    targets = [(12700, 0.010), (11500, 0.019)]
    extra = [(10000, 0.010), (10000, 0.020), (11000, 0.010)]
    targets.extend(extra)

    for L_, S in targets:
        key = (L_, round(S, 3))
        in_opt = key in ref_opt
        in_test = key in ref_test
        if not in_opt and not in_test:
            print(f"\n(L={L_}, S={S}) not in either reference -- skipped")
            continue
        t1 = time.time()

        res = run_strategy(H, L, C, chn_len=L_, stp_pct=S, pv=PV, slpg=SLPG,
                           bars_back=BARS_BACK, e0=E0)
        wis = res.window_stats(is_lo, is_hi)
        wos = res.window_stats(os_lo, os_hi)

        got_is = (wis.profit, wis.worst_dd, wis.stdev, wis.trades)
        got_os = (wos.profit, wos.worst_dd, wos.stdev, wos.trades)

        print(f"\nL={L_}, S={S}   ({time.time()-t1:.2f}s)")
        if in_opt:
            print("vs HO Optimization.xlsx")
            check("IS", got_is, ref_opt[key]["is"])
            check("OS", got_os, ref_opt[key]["os"])
        if in_test:
            print("vs HO MatLab Test.xlsx (IS only; OS window differs)")
            check("IS", got_is, ref_test[key]["is"])

    print(f"\nTotal: {time.time()-t0:.1f}s")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
